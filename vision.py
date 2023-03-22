#! /usr/bin/python3
from imutils.video import VideoStream
from std_msgs.msg import String
from imutils.video import FPS
import face_recognition
import imutils
import pickle
import time
import cv2
import requests
import openpyxl
import time
from datetime import datetime
import rospy
sus = cv2.imread("/home/pi/catkin_ws/src/design2/imposter.jpeg")
sus = imutils.resize(sus, width=800)
notsus = cv2.imread("/home/pi/catkin_ws/src/design2/not imposter.jpg")
notsus = imutils.resize(notsus, width=800)

wb = openpyxl.Workbook()
ws = wb.active
sheet1 = ws
qr_list = []
ws['A1'] = "Code"
ws['B1'] = "ID"
ws['C1'] = "Name"
ws['D1'] = "Facial Reading"
coursecode = 'MCT-111'
window1 = "Scan your face"
window2 = "Qr code detector"
window3 = "Very sus indeed"
window4 = "Not sus"
cv2.namedWindow(window1)
cv2.namedWindow(window2)
cv2.moveWindow(window1, 1500,0)
cv2.moveWindow(window2, 1500,500)

currentname = "unknown"
encodingsP = "/home/pi/catkin_ws/src/design2/encodings.pickle"
cascade = "/home/pi/catkin_ws/src/design2/haarcascade_frontalface_default.xml"
print("[INFO] loading encodings + face detector...")
data = pickle.loads(open(encodingsP, "rb").read())
detector = cv2.CascadeClassifier(cascade)
oldqr = 'nothing'
print("[INFO] starting video stream...")
vs = VideoStream(src=0).start()
time.sleep(2.0)
det = cv2.QRCodeDetector()
row = 2
iterate_row = 0
arduino_data = "null"

def main():
    rospy.init_node('vision')
    rospy.Subscriber("team38", String, callback)
    rospy.spin()
    
def callback(data):
    arduino_data = data.data
    rospy.init_node('vision')
    rospy.Subscriber("team38", String, callback)
    print(arduino_data)
    img = vs.read()
    qrdata, bbox, _ = det.detectAndDecode(img)
    img = imutils.resize(img, width=400)
    if(bbox is not None):
        #for i in range(len(bbox)):
            #cv2.line(img, tuple(bbox[i][0]), tuple(bbox[(i+1) % len(bbox)][0]), color=(255,0, 255), thickness=2)
        cv2.putText(img, qrdata, (int(bbox[0][0][0]), int(bbox[0][0][1]) - 10), cv2.FONT_HERSHEY_SIMPLEX,0.5, (0, 255, 0), 2)
        if qrdata:
                if qrdata != oldqr:
                    print("data found: ", qrdata)
                    oldqr = qrdata
                    qr_list = qrdata.split("  ")
                    if (qr_list[2] != currentname):
                        print("incorrect QR code reason: Face and QR code don't match")
                        cv2.namedWindow(window3)
                        cv2.moveWindow(window3, 600,100)
                        cv2.imshow(window3, sus)
                        
                        
                    elif (qr_list[0] != coursecode):
                        print("incorrect QR code reason: different course code")
                        
                        
                    elif (qr_list[0] == coursecode):
                        cell = sheet1.cell(row=row, column=1)
                        cell.value = qr_list[0]
                        cell = sheet1.cell(row=row, column=2)
                        cell.value = qr_list[1]
                        cell = sheet1.cell(row=row, column=3)
                        cell.value = qr_list[2]
                        wb.save('/home/pi/Desktop/Attendance.xlsx')
                        row = row + 1
                        cv2.namedWindow(window4)
                        cv2.moveWindow(window4, 600,100)
                        cv2.imshow(window4, notsus)
                       
                
    frame = vs.read()
    frame = imutils.resize(frame, width=400)
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    rects = detector.detectMultiScale(gray, scaleFactor=1.1, 
        minNeighbors=5, minSize=(30, 30),
        flags=cv2.CASCADE_SCALE_IMAGE)

    boxes = [(y, x + w, y + h, x) for (x, y, w, h) in rects]

    encodings = face_recognition.face_encodings(rgb, boxes)
    names = []

    for encoding in encodings:

        matches = face_recognition.compare_faces(data["encodings"],
            encoding)
        name = "unknown"

        if True in matches:

            matchedIdxs = [i for (i, b) in enumerate(matches) if b]
            counts = {}

            for i in matchedIdxs:
                name = data["names"][i]
                counts[name] = counts.get(name, 0) + 1


            name = max(counts, key=counts.get)
            if currentname != name:
                cv2. destroyWindow(window4)
                currentname = name
                print("Your attendance has been registered",currentname)
                cell = sheet1.cell(row=row, column=4)
                cell.value = currentname
                wb.save('/home/pi/Desktop/Attendance.xlsx')
                
                
                
        names.append(name)

    for ((top, right, bottom, left), name) in zip(boxes, names):
        cv2.rectangle(frame, (left, top), (right, bottom),
            (0, 255, 225), 2)
        y = top - 15 if top - 15 > 15 else top + 15
        cv2.putText(frame, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX,
            .8, (0, 255, 255), 2)

    # display the image to our screen
    cv2.imshow(window1, frame)
    cv2.imshow(window2, img)
    key = cv2.waitKey(1) & 0xFF
    if key == ord("q") or arduino_data == "off":
        print("The robot has finished its path thank you for using NAZIR")
        exit()
        
    
    


cv2.destroyAllWindows()
vs.stop()
